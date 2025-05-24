import os
import openpyxl # Added import

def parse_scv_init(file_path: str) -> dict:
    """
    Parses a SCV initialization file and returns a dictionary of key-value pairs.

    Args:
        file_path: The path to the SCV initialization file.

    Returns:
        A dictionary containing the parsed key-value pairs.
        Numerical values are converted to floats, others are kept as strings.
        Returns an empty dictionary if the file is not found or an error occurs.
    """
    config_data = {}
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return config_data

    try:
        with open(file_path, 'r') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue

                parts = line.split(None, 1)  # Split by any whitespace, max 1 split
                if len(parts) == 2:
                    key, value_str = parts
                    try:
                        value = float(value_str)
                    except ValueError:
                        value = value_str  # Keep as string if not a float
                    config_data[key] = value
                else:
                    print(f"Warning: Could not parse line: {line}")
    except IOError:
        print(f"Error: Could not read file at {file_path}")
    
    return config_data

def parse_dose_monitor_excel(file_path: str, num_log_files: int) -> list[int]:
    """
    Parses dose and monitor unit data from column "T" of "Sheet1" in an Excel file.

    Args:
        file_path: The path to the Excel file.
        num_log_files: The expected number of log files (rows to read).

    Returns:
        A list of integers from column "T".

    Raises:
        FileNotFoundError: If the Excel file is not found.
        KeyError: If "Sheet1" is not found in the Excel file.
        ValueError: If values in column "T" are not integers, or if the check
                    on row `num_log_files + 1` fails (non-zero numeric value found).
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Error: Excel file not found at {file_path}")

    try:
        workbook = openpyxl.load_workbook(file_path)
    except Exception as e: # Catching general openpyxl load errors
        raise IOError(f"Error: Could not open or read Excel file at {file_path}. Original error: {e}")


    if "Sheet1" not in workbook.sheetnames:
        raise KeyError("Error: 'Sheet1' not found in the Excel file.")
    
    sheet = workbook["Sheet1"]
    
    monitor_units = []
    
    for row_num in range(1, num_log_files + 1):
        cell_value = sheet.cell(row=row_num, column=20).value # Column T is the 20th column
        if cell_value is None:
            # Or handle as error, depending on requirements for missing values within the range
            raise ValueError(f"Error: Empty cell found at T{row_num} within the expected range of {num_log_files} log files.")
        try:
            monitor_units.append(int(cell_value))
        except (ValueError, TypeError):
            raise ValueError(f"Error: Non-integer value '{cell_value}' found in column T at row {row_num}.")
            
    # Check the cell at row num_log_files + 1 in column T
    check_cell_value = sheet.cell(row=num_log_files + 1, column=20).value
    if check_cell_value is not None:
        is_numeric_and_intolerable = False
        try:
            # Attempt to convert to float. If this fails, it's not a numeric value.
            numeric_check_value = float(check_cell_value)
            
            if numeric_check_value != 0.0: # Explicitly compare with 0.0
                is_numeric_and_intolerable = True
        except (ValueError, TypeError):
            # Cell content is not convertible to a float (e.g., text like "N/A").
            # This means it's not a "numeric value that is not zero", so this specific check passes.
            pass
        
        if is_numeric_and_intolerable:
            raise ValueError(
                f"Dose monitor range file length does not fit the number of log files. "
                f"Found non-zero value '{check_cell_value}' at T{num_log_files + 1}."
            )
            
    return monitor_units

if __name__ == '__main__':
    # --- Test parse_scv_init ---
    dummy_scv_content = """
# INITIALIZE FILE
XPRESETOFFSET	32767
YPRESETOFFSET	32767
XPRESETGAIN	0.0054705
TIMEGAIN	0.0600745
RECFILEDUMP	YES
"""
    dummy_scv_path = "dummy_scv_init.txt"
    with open(dummy_scv_path, 'w') as f:
        f.write(dummy_scv_content)

    print(f"--- Testing parse_scv_init ---")
    print(f"Testing with valid file: {dummy_scv_path}")
    parsed_data = parse_scv_init(dummy_scv_path)
    print("Parsed data:")
    for key, value in parsed_data.items():
        print(f"  {key}: {value} (type: {type(value)})")
    
    print("\nTesting with non-existent file:")
    parsed_data_non_existent = parse_scv_init("non_existent_scv_file.txt")
    print(f"Parsed data for non-existent file: {parsed_data_non_existent}")
    os.remove(dummy_scv_path)
    print("--- End Test parse_scv_init ---\n")

    # --- Test parse_dose_monitor_excel ---
    print("--- Testing parse_dose_monitor_excel ---")

    # Helper to create dummy excel files
    def create_dummy_excel(file_name, data_sheet1_col_T):
        wb = openpyxl.Workbook()
        if "Sheet1" in wb.sheetnames:
            sheet = wb["Sheet1"]
        else: # Should not happen with new workbook, but good practice
            sheet = wb.active
            sheet.title = "Sheet1"
        
        for i, val in enumerate(data_sheet1_col_T):
            sheet.cell(row=i + 1, column=20).value = val # Column T
        wb.save(file_name)
        print(f"Created dummy Excel: {file_name}")

    # Test Case 1: Valid file, correct data
    valid_excel_path = "dummy_excel_valid.xlsx"
    create_dummy_excel(valid_excel_path, [10, 20, 30, 0]) # T4 is 0
    try:
        print("\nTest Case 1: Valid file, correct data (num_log_files=3)")
        monitor_units = parse_dose_monitor_excel(valid_excel_path, 3)
        print(f"Parsed monitor units: {monitor_units}")
        assert monitor_units == [10, 20, 30]
    except Exception as e:
        print(f"Error in Test Case 1: {e}")
    os.remove(valid_excel_path)

    # Test Case 2: Valid file, T(num_log_files+1) is non-zero
    invalid_excel_path_check = "dummy_excel_invalid_check.xlsx"
    create_dummy_excel(invalid_excel_path_check, [10, 20, 30, 5]) # T4 is 5
    try:
        print("\nTest Case 2: Valid file, T(num_log_files+1) is non-zero (num_log_files=3)")
        parse_dose_monitor_excel(invalid_excel_path_check, 3)
        print("Error: Expected ValueError was not raised.")
    except ValueError as e:
        print(f"Caught expected error: {e}")
        assert "Dose monitor range file length does not fit" in str(e)
    except Exception as e:
        print(f"Error in Test Case 2 (unexpected exception): {e}")
    os.remove(invalid_excel_path_check)

    # Test Case 3: File not found
    try:
        print("\nTest Case 3: File not found")
        parse_dose_monitor_excel("non_existent_excel.xlsx", 3)
        print("Error: Expected FileNotFoundError was not raised.")
    except FileNotFoundError as e:
        print(f"Caught expected error: {e}")
    except Exception as e:
        print(f"Error in Test Case 3 (unexpected exception): {e}")

    # Test Case 4: Sheet1 not found
    no_sheet1_excel_path = "dummy_excel_no_sheet1.xlsx"
    wb_no_sheet1 = openpyxl.Workbook()
    wb_no_sheet1.active.title = "NotSheet1"
    wb_no_sheet1.save(no_sheet1_excel_path)
    print(f"Created dummy Excel: {no_sheet1_excel_path}")
    try:
        print("\nTest Case 4: Sheet1 not found")
        parse_dose_monitor_excel(no_sheet1_excel_path, 3)
        print("Error: Expected KeyError was not raised.")
    except KeyError as e:
        print(f"Caught expected error: {e}")
    except Exception as e:
        print(f"Error in Test Case 4 (unexpected exception): {e}")
    os.remove(no_sheet1_excel_path)

    # Test Case 5: Non-integer value in column T
    non_int_excel_path = "dummy_excel_non_int.xlsx"
    create_dummy_excel(non_int_excel_path, [10, "not_an_int", 30])
    try:
        print("\nTest Case 5: Non-integer value in column T")
        parse_dose_monitor_excel(non_int_excel_path, 3)
        print("Error: Expected ValueError was not raised.")
    except ValueError as e:
        print(f"Caught expected error: {e}")
        assert "Non-integer value" in str(e)
    except Exception as e:
        print(f"Error in Test Case 5 (unexpected exception): {e}")
    os.remove(non_int_excel_path)

    # Test Case 6: Empty cell within expected range
    empty_cell_excel_path = "dummy_excel_empty_cell.xlsx"
    create_dummy_excel(empty_cell_excel_path, [10, None, 30])
    try:
        print("\nTest Case 6: Empty cell in column T within range")
        parse_dose_monitor_excel(empty_cell_excel_path, 3)
        print("Error: Expected ValueError was not raised.")
    except ValueError as e:
        print(f"Caught expected error: {e}")
        assert "Empty cell found" in str(e)
    except Exception as e:
        print(f"Error in Test Case 6 (unexpected exception): {e}")
    os.remove(empty_cell_excel_path)
    
    # Test Case 7: Valid file, T(num_log_files+1) is empty (should pass)
    valid_excel_empty_check_path = "dummy_excel_valid_empty_check.xlsx"
    create_dummy_excel(valid_excel_empty_check_path, [10, 20, 30, None]) # T4 is None
    try:
        print("\nTest Case 7: Valid file, T(num_log_files+1) is empty (num_log_files=3)")
        monitor_units = parse_dose_monitor_excel(valid_excel_empty_check_path, 3)
        print(f"Parsed monitor units: {monitor_units}")
        assert monitor_units == [10, 20, 30]
    except Exception as e:
        print(f"Error in Test Case 7: {e}")
    os.remove(valid_excel_empty_check_path)

    print("--- End Test parse_dose_monitor_excel ---")
